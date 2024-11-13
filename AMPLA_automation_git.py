from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import os
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
import re
from openai import OpenAI
from docx import Document
import streamlit as st

#-----LOG-IN INFO: DELETE LATER
# roberto.vieyra-hernandez@dharmann.com
# 2701113iV 

# ------- Set up WebDriver
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

download_dir = r"C:\Users\cloea\Desktop\Resumes" # Download directory for resumes

chrome_options.add_experimental_option('prefs', {
    "download.default_directory": download_dir,  # Set the download directory
    "download.prompt_for_download": False,  # Disable download prompt
    "plugins.always_open_pdf_externally": True  # Prevent PDF viewer from opening
})

driver = webdriver.Chrome(options=chrome_options)

# ------- Open the URL
url = "https://dharmannstudios.bamboohr.com/hiring/jobs"
driver.get(url)
driver.maximize_window()

# ------- Log-in and go to 'Job Openings'
print("Waiting for log-in credentials...")

# Wait to log-in
WebDriverWait(driver, 120).until(EC.url_changes(url))
WebDriverWait(driver, 30).until(EC.url_to_be(url))

# ------- Read the Excel file to access the positions you want to filter
df = pd.read_excel(r"C:\Users\cloea\Desktop\Job Openings.xlsx")
positions = df['Position'].tolist()  # List of positions to search for
position_count = 0

# ------- Iterate through each position
for position in positions:
    
    position_count+=1
    print("Position: ", position)

    found = False

    while True:
        print("Checking current page...")

        # Wait until rows are present inside tbody
        rows = WebDriverWait(driver, 30).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "tr"))
        )

        # Loop through each row to search for the specific position
        for row in rows:
            # Get all td elements in the row
            tds = row.find_elements(By.TAG_NAME, "td")

            # Ensure there are at least two td elements
            if len(tds) > 1:
                # Find the <a> element inside the second <td> element (job opening column)
                try:
                    a_element = tds[1].find_element(By.TAG_NAME, "a")
                    # Get the text inside the <a> element and check if it matches the target name
                    if a_element.text == position:
                        print(f"Found '{position}' in the table.")
                        
                        # Click the <a> element
                        WebDriverWait(driver, 30).until(EC.element_to_be_clickable(a_element))
                        driver.execute_script("arguments[0].click();", a_element)
                        
                        found = True
                        break  # Exit the row loop if position is found
                except Exception as e:
                    continue  # Skip to the next row if <a> element is not found

        if found:
            break  # Exit the page loop if position is found

        # Attempt to find the "Next" button to move to the next page
        try:
            next_button = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Next')]"))  # Adjust the XPath as needed
            )
            next_button.click()  # Click on the next page
            time.sleep(1)  # Add a brief wait for the next page to load
        except Exception as e:
            break  # Exit if no more pages are available

    if not found:
        print(f"'{position}' not found in the table.")

    else: # Filter new candidates
        
        # Wait for the toggle button to be present and click it to show the options
        toggle_button = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CLASS_NAME, "fab-SelectToggle__toggleButton")))
        time.sleep(1) 
        toggle_button.click()  # Click to open the menu
        time.sleep(1) 

        # Locate the option you want to select
        option_to_select = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//div[text()='New']")))
        option_to_select.click()  # Click the option
        time.sleep(1) 

        print("New candidates filtered")

        time.sleep(1)

        #----- Check if there's new candidates

        # Locate the <h4> element and extract its text
        h4_element = driver.find_element(By.TAG_NAME, 'h4')  # Locate by tag name
        h4_element = h4_element.text

        # The regular expression r'\d+' matches the first occurrence of one or more digits (\d+). re.search() finds the first match and returns a match object.
        new_candidates_num  = re.search(r'\d+', h4_element)
        new_candidates_num = int(new_candidates_num.group()) #convert the match (substring) to int

        if new_candidates_num == 0:
            print('No new candidates found') 
        
        #-------Download the cv from each candidate and iterate through each position
        else:
            print(new_candidates_num,' candidate(s) found')

            candidate_rows = new_candidates_num+1 #actual number of candidates in that specific position

            for i in range(candidate_rows):
                
                # Wait until rows are present inside tbody
                candidate= WebDriverWait(driver, 30).until(
                    EC.presence_of_all_elements_located((By.TAG_NAME, "tr"))
                )

                # Get all td elements in the row 
                td_candidate = candidate[i].find_elements(By.TAG_NAME, "td")

                try:
                #         # Look for <a> element in the second td element (td_candidate[1])---> candidate info
                    a_element_candidate = td_candidate[1].find_element(By.TAG_NAME, "a")

                    # Get the text inside the <a> element (candidate's name)
                    candidate_name = a_element_candidate.text

                    if a_element_candidate: #Candidate's name found
                        print("Candidate ",i ,f": '{candidate_name}'")

                        # Click the <a> element (candidate's name)
                        WebDriverWait(driver, 30).until(EC.element_to_be_clickable(a_element_candidate))
                        driver.execute_script("arguments[0].click();",a_element_candidate)

                        time.sleep(2)

                        # Locate and click the PDF link
                        download_button =  WebDriverWait(driver, 30).until(
                                EC.element_to_be_clickable((By.ID,'download'))  # Adjust the selector as needed
                        )
                        download_button.click()

                        # Wait for the download to complete 
                        time.sleep(5)  # Wait for 5 seconds or adjust based on file size

                        # Check for the most recently downloaded file
                        files = os.listdir(download_dir)  # List all files in the download directory
                        full_paths = [os.path.join(download_dir, f) for f in files]  # Create full paths

                        # Get the most recently downloaded file (path)
                        most_recent_file = max(full_paths, key=os.path.getctime)
                        print(most_recent_file)

                        # Initialize an empty string to hold the extracted text
                        all_text = " "

                        #Check if the resume is a pdf 
                        if most_recent_file.endswith(".pdf"):
                            # Attempt to open the PDF file
                            with pdfplumber.open(most_recent_file) as pdf:
                                
                                # Loop through the pages of the PDF
                                for page in pdf.pages:
                                    # Extract text from each page
                                    text = page.extract_text()
                                    if text:
                                        all_text += text

                        #Check if the resume is a Word document 
                        elif most_recent_file.endswith((".doc", ".docx")):
                            # Load the document
                            doc = Document(most_recent_file)
                            
                            # Extract text from each paragraph
                            all_text = "\n".join([para.text for para in doc.paragraphs])
                            
                        else:
                            print("The resume was not uploaded on a valid format")
                            all_text = "no resume provided"

                        # Read the Excel file
                        df = pd.read_excel(r"C:\Users\cloea\Desktop\Job Openings.xlsx")
                        df_row = position_count - 1

                        # Extract value of 'Keywords' column where 'Position' is 'Senior Producer - YouTube'
                        job_description = df.iat[df_row, 1] 

                        #----Open AI model

                        # Set your OpenAI API key
                        # Access the secret API key
                        api_key = st.secrets["OPENAI_API_KEY"]

                        # Create the AsyncOpenAI client
                        client = OpenAI(api_key=api_key)

                        response = client.chat.completions.create(
                            model="gpt-4o",
                            messages=[
                            {
                                "role": "system",
                                "content": "You will be provided with a job description and a candidate's resume. Please analyze the candidate's qualifications, skills, and experience in relation to the job description. Rate the candidate on a scale of 0 to 5 based on their overall fit for the position, write the rating in the format: 'Rating : x/5'. Include a brief explanation for the rating, focusing on how well the candidate's experience and skills match the key requirements of the job. Answer in bullet format. "
                            },
                            {
                                "role": "user",
                                "content": "Job description: " + f"{job_description}" + "The candidate's resume is: " + f"{all_text}"
                            }
                            ],
                            temperature=0.2, #less random, more deterministic
                            max_tokens=1000,
                            top_p=1,
                            n = 1, #only one answer per prompt - no variation
                        )

                        answer = response.choices[0].message.content

                        print(answer)

                        target_word = "Rating: "

                        # Find the starting and ending index of the target word
                        start_index = answer.find(target_word)
                        end_index = start_index + len(target_word)

                        # Check if the word was found
                        if start_index != -1:
                            # Get the character after the word (if any)
                            if end_index < len(answer):
                                rating = int(answer[end_index])
                        else:
                            print(f"'{target_word}' not found in the answer.")


                        # Load the existing Excel file
                        file_path =(r"C:\Users\cloea\Desktop\Job Openings.xlsx")
                        sheet_name = position

                        # Create a sample DataFrame
                        data = {
                            'Name': [None],
                            'Rating': [None],
                            'Feedback': [None]
                        }

                        df2 = pd.DataFrame(data)

                        df2.loc[i-1, 'Name'] = candidate_name 
                        df2.loc[i-1, 'Rating'] = rating
                        df2.loc[i-1, 'Feedback'] = answer

                        # Load the existing workbook if it exists
                        try:
                            # Try loading the existing Excel file
                            book = load_workbook(file_path)
                            
                            # Check if the sheet exists
                            if sheet_name in book.sheetnames:
                                print(f"Sheet '{sheet_name}' already exists. Updating it...")
                                
                                # If the sheet exists, load it into a DataFrame
                                existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                                row_count = len(existing_df)

                                # Check if the name exists in the first column 
                                name_exists = candidate_name  in existing_df.iloc[:, 0].values #boolean value

                                if name_exists:
                                    print(candidate_name, " was found on the database: Updating rating...")
                                    # Update the second column (rating) where the name is found
                                    existing_df.loc[existing_df.iloc[:, 0] == candidate_name, existing_df.columns[1]] = rating
                                    # Update the third column (answer) where the name is found
                                    existing_df.loc[existing_df.iloc[:, 0] == candidate_name, existing_df.columns[2]] = answer

                                else:
                                    # Update the DataFrame: Add a new row at the end (row_count value)
                                    existing_df.loc[row_count, 'Name'] = candidate_name 
                                    existing_df.loc[row_count, 'Rating'] = rating  
                                    existing_df.loc[row_count, 'Feedback'] = answer

                                # Save back to the same sheet using ExcelWriter in append mode
                                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                    existing_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            else:
                                print(f"Creating a new sheet named '{sheet_name}'...")
                                
                                # If the sheet does not exist, write the DataFrame to a new sheet
                                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                    df2.to_excel(writer, sheet_name=sheet_name, index=False)

                            # Adjust column width, wrap text, and add borders
                            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                # Access the workbook and the specific sheet
                                workbook = writer.book
                                worksheet = workbook[sheet_name]
                                col_num = 0

                                # Define border style
                                thin_border = Border(left=Side(style='thin'), 
                                                    right=Side(style='thin'), 
                                                    top=Side(style='thin'), 
                                                    bottom=Side(style='thin'))

                                # Adjust the width of specific columns, enable wrap text, and add borders
                                for col in worksheet.columns:
                                    max_length = 0
                                    column = col[0].column_letter  # Get the column letter

                                    if col_num  == 0 or col_num  == 1:    
                                        adjusted_width = 10
                                    else:
                                        adjusted_width = 125

                                    worksheet.column_dimensions[column].width = adjusted_width # Set the column width
                                    col_num += 1

                                    for cell in col:
                                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                                        cell.border = thin_border  # Apply the border
                                    

                            # Save the workbook
                            workbook.save(file_path)

                        except FileNotFoundError:
                            print("File does not exist.")

                        print("Database updated: Operation completed successfully!")

                        #------ Delete the file (resume)
                        
                        # Check if the file exists
                        if os.path.exists(most_recent_file):
                            os.remove(most_recent_file)  # Delete the file
                            print(f"Deleted file: {most_recent_file}")
                        else:
                            print("The specified file does not exist.")

                        time.sleep(2)

                        #------ Add stars to each candidate depending on the rating 
                        stars = rating
                        
                        match stars:
                            case 1:
                                rating_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                                (By.XPATH, "/html/body/div[3]/div[3]/main/div/div/div[2]/div/section/div/div[2]/div/aside/div[1]/section/div[1]/div/div/div/div[1]")
                                ))
                            case 2:
                                rating_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                                (By.XPATH, "/html/body/div[3]/div[3]/main/div/div/div[2]/div/section/div/div[2]/div/aside/div[1]/section/div[1]/div/div/div/div[2]")
                                ))
                            case 3:
                                    rating_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                                (By.XPATH, "/html/body/div[3]/div[3]/main/div/div/div[2]/div/section/div/div[2]/div/aside/div[1]/section/div[1]/div/div/div/div[3]")
                                ))
                            case 4:
                                    rating_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                                (By.XPATH, "/html/body/div[3]/div[3]/main/div/div/div[2]/div/section/div/div[2]/div/aside/div[1]/section/div[1]/div/div/div/div[4]")
                                ))
                            case 5:
                                    rating_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                                (By.XPATH, "/html/body/div[3]/div[3]/main/div/div/div[2]/div/section/div/div[2]/div/aside/div[1]/section/div[1]/div/div/div/div[5]")
                                ))

                        # Click the element
                        rating_element.click()
                        print("*****Candidate rated on the plataform******")
                        time.sleep(2)
                        
                        # Go to the previous page (candidate list)
                        driver.back()
                        time.sleep(2)

                except Exception as e:
                    continue  # Skip to the next candidate_row if <a> element (candidate) is not found 
    
    time.sleep(3)

print("Done")

time.sleep(5)

driver.quit()
